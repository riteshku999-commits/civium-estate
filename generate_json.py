#!/usr/bin/env python3
"""
generate_json.py  —  Civium Estate Property Manager
====================================================
Run this script after editing the Excel file to:
  1. Read all property rows from 'Properties' sheet
  2. Write data/properties.json  (ready to push to GitHub)
  3. Refresh the 'JSON Preview' sheet in the Excel itself

Usage:
    python generate_json.py

The script automatically finds the Excel file in the same folder.
It looks for:  Civium_Estate_Property_Manager.xlsx

Requirements:
    pip install openpyxl

"""

import json
import os
import sys
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: openpyxl is not installed.")
    print("Run:  pip install openpyxl")
    sys.exit(1)


# ─── Configuration ────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent
EXCEL_FILE   = SCRIPT_DIR / "Civium_Estate_Property_Manager.xlsx"
OUTPUT_JSON  = SCRIPT_DIR / "data" / "properties.json"

# Column header → field mapping (must match Excel headers exactly)
SHEET_NAME   = "Properties"
DATA_START   = 5   # Row where property data begins


# ─── Column header → internal key mapping ────────────────────
COL_MAP = {
    "ID":                 "id",
    "Title":              "title",
    "Description":        "description",
    "Category":           "category",
    "Type":               "type",
    "Location":           "location",
    "Price (Lakhs)":      "price",
    "Price Label":        "priceLabel",
    "Old Price Label":    "oldPriceLabel",
    "Badge Text":         "badge",
    "Badge Class":        "badgeClass",
    "Is Hot Deal":        "isHotDeal",
    "Featured":           "featured",
    "Sold Out":           "soldOut",
    "Active":             "active",
    "Image 1 URL":        "img1",
    "Image 2 URL":        "img2",
    "Image 3 URL":        "img3",
    "Image 4 URL":        "img4",
    "Image 5 URL":        "img5",
    "Video URL":          "video",
    "Facing":             "facing",
    "Vastu":              "vastu",
    "Floor":              "floor",
    "Total Floors":       "totalFloors",
    "Parking":            "parking",
    "Furnished":          "furnished",
    "Age of Property":    "ageOfProperty",
    "Possession":         "possession",
    "Power Backup":       "powerBackup",
    "Metro Name":         "metroName",
    "Metro Dist (km)":    "metroDist",
    "Railway Name":       "railwayName",
    "Railway Dist (km)":  "railwayDist",
    "Hospital Name":      "hospitalName",
    "Hospital Dist (km)": "hospitalDist",
    "School Name":        "schoolName",
    "School Dist (km)":   "schoolDist",
    "Airport Name":       "airportName",
    "Airport Dist (km)":  "airportDist",
    "Map Link":           "mapLink",
}

BOOL_FIELDS = {"isHotDeal", "featured", "soldOut", "active", "vastu"}


def cell_val(cell):
    """Return stripped string value of a cell, or empty string."""
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def to_bool(val):
    return str(val).upper() == "YES"


def to_num(val):
    """Convert to int or float, or return original string."""
    if val == "":
        return None
    try:
        f = float(val)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return val


def build_property(row_data):
    """Convert a dict of {field_key: value} into the JSON property structure."""
    p = {}

    # ── Core fields ────────────────────────────────────────────
    p["id"]          = row_data.get("id", "")
    p["title"]       = row_data.get("title", "")
    p["description"] = row_data.get("description", "")
    p["category"]    = row_data.get("category", "")
    p["isHotDeal"]   = to_bool(row_data.get("isHotDeal", "NO"))
    p["featured"]    = to_bool(row_data.get("featured", "NO"))
    p["badge"]       = row_data.get("badge", "")
    p["badgeClass"]  = row_data.get("badgeClass", "")

    price_raw = row_data.get("price", "")
    p["price"] = to_num(price_raw) if price_raw else 0

    p["priceLabel"]  = row_data.get("priceLabel", "")

    old_price = row_data.get("oldPriceLabel", "")
    if old_price:
        p["oldPriceLabel"] = old_price

    p["location"]    = row_data.get("location", "")
    p["type"]        = row_data.get("type", "")
    p["soldOut"]     = to_bool(row_data.get("soldOut", "NO"))
    p["active"]      = to_bool(row_data.get("active", "YES"))

    # ── Media ──────────────────────────────────────────────────
    media = []
    for key in ["img1", "img2", "img3", "img4", "img5"]:
        url = row_data.get(key, "")
        if url:
            media.append({"type": "image", "src": url})
    video = row_data.get("video", "")
    if video:
        media.append({"type": "video", "src": video})
    p["media"] = media

    # ── Attributes ─────────────────────────────────────────────
    attrs = {}

    if row_data.get("facing"):
        attrs["facing"] = row_data["facing"]

    vastu_raw = row_data.get("vastu", "")
    if vastu_raw.upper() in ("YES", "NO"):
        attrs["vastu"] = (vastu_raw.upper() == "YES")

    floor_raw = row_data.get("floor", "")
    if floor_raw:
        attrs["floor"] = to_num(floor_raw)

    total_floors = row_data.get("totalFloors", "")
    if total_floors:
        attrs["totalFloors"] = to_num(total_floors)

    if row_data.get("parking"):
        attrs["parking"] = row_data["parking"]

    if row_data.get("furnished"):
        attrs["furnished"] = row_data["furnished"]

    if row_data.get("ageOfProperty"):
        attrs["ageOfProperty"] = row_data["ageOfProperty"]

    if row_data.get("possession"):
        attrs["possession"] = row_data["possession"]

    if row_data.get("powerBackup"):
        attrs["powerBackup"] = row_data["powerBackup"]

    # Proximity attrs
    for prefix, key in [
        ("metro",    "nearestMetro"),
        ("railway",  "nearestRailway"),
        ("hospital", "nearestHospital"),
        ("school",   "nearestSchool"),
        ("airport",  "nearestAirport"),
    ]:
        name = row_data.get(f"{prefix}Name", "")
        dist = row_data.get(f"{prefix}Dist", "")
        if name:
            attrs[key] = {
                "name": name,
                "distanceKm": to_num(dist) if dist else 0
            }

    if row_data.get("mapLink"):
        attrs["mapLink"] = row_data["mapLink"]

    p["attributes"] = attrs
    return p


def read_excel():
    """Read the Excel and return list of property dicts."""
    if not EXCEL_FILE.exists():
        print(f"ERROR: Cannot find Excel file at:  {EXCEL_FILE}")
        print("Make sure generate_json.py is in the same folder as the Excel file.")
        sys.exit(1)

    print(f"Reading: {EXCEL_FILE.name}")
    wb = load_workbook(str(EXCEL_FILE), data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found in Excel file.")
        sys.exit(1)

    ws = wb[SHEET_NAME]

    # Find column positions from header row (row 4)
    col_positions = {}
    for cell in ws[4]:
        hdr = str(cell.value or "").strip()
        if hdr in COL_MAP:
            col_positions[COL_MAP[hdr]] = cell.column

    found_keys = set(col_positions.keys())
    expected_keys = set(COL_MAP.values())
    missing_keys = expected_keys - found_keys
    if missing_keys:
        missing_hdrs = [h for h, k in COL_MAP.items() if k in missing_keys]
        print(f"WARNING: Could not find these columns in Excel header row 4: {missing_hdrs}")

    # Read data rows
    properties = []
    for row_num in range(DATA_START, ws.max_row + 1):
        row_data = {}
        for field_key, col_idx in col_positions.items():
            row_data[field_key] = cell_val(ws.cell(row=row_num, column=col_idx))

        # Skip blank rows (no ID)
        if not row_data.get("id"):
            continue

        prop = build_property(row_data)
        properties.append(prop)

    wb.close()
    return properties


def write_json(properties):
    """Write properties.json."""
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    output = {"properties": properties}
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    print(f"Written: {OUTPUT_JSON}  ({len(properties)} properties)")


def refresh_preview_sheet(properties):
    """Rewrite the JSON Preview sheet in the Excel with the fresh JSON."""
    try:
        wb = load_workbook(str(EXCEL_FILE))

        preview_name = "JSON Preview"
        if preview_name in wb.sheetnames:
            del wb[preview_name]

        ws2 = wb.create_sheet(preview_name, 1)  # insert after Properties
        ws2.sheet_view.showGridLines = False
        ws2.column_dimensions["A"].width = 3
        ws2.column_dimensions["B"].width = 185

        def fill(h): return PatternFill("solid", start_color=h, fgColor=h)

        ws2.merge_cells("A1:B1")
        c = ws2["A1"]
        c.value = "JSON Preview  -  Current properties.json content"
        c.fill  = fill("0D0D0D")
        c.font  = Font(bold=True, size=14, color="C5A47E", name="Arial")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 36

        ws2.merge_cells("A2:B2")
        c = ws2["A2"]
        c.value = f"Last updated by generate_json.py  |  {len(properties)} properties loaded"
        c.fill  = fill("1A1A1A")
        c.font  = Font(size=10, color="90EE90", italic=True, name="Arial")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[2].height = 24

        ws2.merge_cells("A3:B3")
        c = ws2["A3"]
        c.value = "  ---  data/properties.json  ---"
        c.fill  = fill("141414")
        c.font  = Font(bold=True, size=10, color="666666", name="Courier New")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[3].height = 18

        json_str = json.dumps({"properties": properties}, indent=2, ensure_ascii=False)

        KEY   = "9CDCFE"; STR  = "CE9178"; BOOL = "569CD6"
        NUM   = "B5CEA8"; PUN  = "D4D4D4"; GLD  = "C5A47E"

        for li, line in enumerate(json_str.split("\n"), start=4):
            cell = ws2.cell(row=li, column=2, value=line)
            s = line.strip()
            if s in ("{", "}", "[", "]", "},") or s.endswith("[") or s.endswith("{"):
                color = GLD
            elif ": true" in s or ": false" in s or ": null" in s:
                color = BOOL
            elif s.startswith('"') and '": "' in s:
                color = STR
            elif s.startswith('"') and '": ' in s:
                try:
                    float(s.split('": ')[1].rstrip(','))
                    color = NUM
                except:
                    color = KEY
            elif s.startswith('"') and s.endswith('"'):
                color = STR
            else:
                color = PUN
            cell.font      = Font(name="Courier New", size=9.5, color=color)
            cell.fill      = fill("1E1E1E")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws2.row_dimensions[li].height = 14

        wb.save(str(EXCEL_FILE))
        print(f"Refreshed: JSON Preview sheet in {EXCEL_FILE.name}")

    except Exception as e:
        print(f"Note: Could not refresh JSON Preview sheet: {e}")
        print("(This is fine — properties.json was still written successfully)")


def main():
    print("=" * 55)
    print("  Civium Estate — Property JSON Generator")
    print("=" * 55)

    properties = read_excel()

    if not properties:
        print("WARNING: No properties found. Check that the Properties sheet has data starting from row 5.")
        sys.exit(1)

    print(f"Found {len(properties)} properties in Excel")

    # Show active vs inactive count
    active   = [p for p in properties if p.get("active")]
    inactive = [p for p in properties if not p.get("active")]
    if inactive:
        print(f"  Active:   {len(active)}")
        print(f"  Hidden:   {len(inactive)}  (Active = NO)")

    write_json(properties)
    refresh_preview_sheet(properties)

    print()
    print("Done! Next steps:")
    print("  1. Commit data/properties.json to GitHub")
    print("  2. Push to GitHub")
    print("  3. Website updates automatically within 30 seconds")
    print()


if __name__ == "__main__":
    main()
